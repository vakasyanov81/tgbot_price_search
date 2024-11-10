import os
from collections import OrderedDict
from typing import NamedTuple, Type

import prettytable as pt
from openpyxl import load_workbook


class PriceInfo(NamedTuple):
    price: float
    price_purchase: float
    rest: str
    row_id: int


price_data = dict[str, list[PriceInfo]]


class PriceListSimple:
    def __init__(
        self, search_driver: Type["PriceSearchInterface"], price_list: price_data
    ):
        self.search_driver = search_driver
        self.price_list = price_list

    def search(self, search_string: str, search_limit=10):
        return self.search_driver(self.price_list).search(search_string, search_limit)


class PriceList:
    def __init__(self, search_driver: Type["PriceSearchInterface"], file_name):
        self.price_list: price_data = OrderedDict()
        self.file_name = file_name
        self.search_driver = search_driver

    def search(self, search_string: str, search_limit=10):
        return self.search_driver(self.price_list).search(search_string, search_limit)

    def load_price(self):
        _file = os.getcwd() + "/" + self.file_name

        try:
            wb = load_workbook(filename=_file)
        except FileNotFoundError:
            return

        sheet = wb["price"]

        for i in range(1, sheet.max_row + 1):
            title = sheet[i][2].value.lower()
            price = sheet[i][6].value  # цена продажи
            price_purchase = sheet[i][5].value  # цена закупочная
            rest = sheet[i][9].value
            if title not in self.price_list:
                self.price_list[title] = []
            self.price_list[title].append(PriceInfo(price, price_purchase, rest, i))

    def clear(self):
        self.price_list = OrderedDict()


class PriceSearchInterface:
    def __init__(self, price_list: price_data):
        self.price_list = price_list

    def search(self, search_string: str, search_limit: int):
        raise NotImplementedError


class PriceListSearch(PriceSearchInterface):

    def search(self, search_string: str, search_limit=10):
        """Поиск по названию позиции в прайсе"""
        result = self.strict_search(
            search_string, search_limit
        ) or self.no_accurate_search(search_string, search_limit)
        return result

    def strict_search(self, search_string: str, search_limit=10):
        """Строгий поиск по точному вхождению"""
        result = dict()
        search_string = search_string.lower()
        search_count = 0

        for title, prices in self.price_list.items():
            if search_string in title:
                result[title] = prices
                search_count += 1
            if search_count >= search_limit:
                break
        return result

    def no_accurate_search(self, search_string: str, search_limit=10):
        """Не строгий поиск"""
        result = dict()
        search_string = search_string.lower()
        search_count = 0
        search_chunks = [s.strip() for s in search_string.split() if len(s.strip())]
        search_chunks_len = len(search_chunks)
        if search_chunks_len == 1:
            return result

        for title, prices in self.price_list.items():
            searched_chunk_count = 0
            for search_chunk in search_chunks:
                if search_chunk in title:
                    searched_chunk_count += 1
            if searched_chunk_count == search_chunks_len:
                result[title] = prices
                search_count += 1
            if search_count >= search_limit:
                break
        return result


def prepare_search_result(result):
    # print(result)
    table = pt.PrettyTable(["название", "цена", "остаток"])
    table.align["название"] = "l"
    table.align["цена"] = "r"
    table.align["остаток"] = "r"
    for title, prices in result.items():
        table.add_row(
            (title, f"{prices[0].price_purchase}/{prices[0].price}", prices[0].rest)
        )
    # response_txt = ''  # find_nomenclature(message.text)
    response_txt = f"```{table}```"
    # print(f"response: {response_txt}")
    return response_txt


def get_instance_price_list(
    price_driver: str, search_driver: str = "PriceListSearch", *args
):
    price_driver_instances = {
        "PriceList": PriceList,
        "PriceListSimple": PriceListSimple,
    }
    search_driver_instance = {"PriceListSearch": PriceListSearch}
    return price_driver_instances.get(price_driver)(
        search_driver_instance.get(search_driver), *args
    )


if __name__ == "__main__":

    def search(title, search_string):
        data = {title: PriceInfo(10, 10, "1", 1)}
        _price_list = get_instance_price_list(
            "PriceListSimple", "PriceListSearch", data
        )
        return _price_list.search("185 65 r15")

    assert {} == search("185/60r15 bfgoodrich g-force winter 2 88t", "185 65 r15")
    assert "185/65r15 bfgoodrich g-force winter 2 88t" in search(
        "185/65r15 bfgoodrich g-force winter 2 88t", "185 65 r15"
    )
