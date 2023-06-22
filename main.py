import glob
import json
import os
from collections import OrderedDict
from enum import Enum
from typing import NamedTuple

import telebot
import openai

from config import Config
from oga import save_oga_to_mp3
from openpyxl import load_workbook
import prettytable as pt

# Autosnab54Info

# Создаем экземпляр бота
bot = telebot.TeleBot(Config.TG_TOKEN)
openai.api_key = Config.OPEN_AI_TOKEN


class Commands(Enum):
    DOWNLOAD_PRICE = 1


class Command:

    def __init__(self):
        self.current_command = {}

    def set_command(self, user_id: int, cmd: Commands):
        if type(cmd) != Commands:
            raise 'Wrong type command!'
        self.current_command[user_id] = cmd

    def current_state(self, user_id):
        return self.current_command.get(user_id)

    def clear(self, user_id):
        self.current_command[user_id] = None


class PriceInfo(NamedTuple):
    price: float
    rest: str
    row_id: int


class PriceList:
    def __init__(self):
        self.price_list: dict[str, list[PriceInfo]] = OrderedDict()

    def load_price(self, file_name):
        _file = os.getcwd() + '/upload/' + file_name

        try:
            wb = load_workbook(filename=_file)
        except FileNotFoundError:
            return

        sheet = wb['price']

        for i in range(1, sheet.max_row + 1):
            title = sheet[i][2].value.lower()
            price = sheet[i][4].value
            rest = sheet[i][7].value
            if title not in self.price_list:
                self.price_list[title] = []
            self.price_list[title].append(PriceInfo(price, rest, i))

    def search(self, search_string: str, search_limit=10):
        result = dict()
        print(f'search: --{search_string}--')
        search_string = search_string.lower()
        search_count = 0

        for title, prices in self.price_list.items():
            if search_string in title:
                result[title] = prices
                search_count += 1
            if search_count >= search_limit:
                break
        return result

    def clear(self):
        self.price_list = OrderedDict()


price_list = PriceList()
price_list.load_price('price.xlsx')


command = Command()


# Обработчик команды /start
@bot.message_handler(commands=['start'])
def send_welcome(message):
    if not check_rights(message.from_user.id):
        replay_not_right_msg(message)
        return
    bot.reply_to(message, 'Привет! Это бот для поиска позиций по прайс листу')


# Обработчик команды /help
@bot.message_handler(commands=['help'])
def send_welcome(message):
    if not check_rights(message.from_user.id):
        replay_not_right_msg(message)
        return
    bot.reply_to(message, 'Для поиска по прайсу необходимо загрузить файл price.xlsx выполнив команду /download')


# Обработчик команды /download
@bot.message_handler(commands=['download'])
def set_download_state(message):
    if not check_rights(message.from_user.id):
        replay_not_right_msg(message)
        return
    command.set_command(message.from_user.id, Commands.DOWNLOAD_PRICE)
    bot.reply_to(message, 'Загрузите файл с остатками price.xlsx')


# Обработчик загрузки файла
@bot.message_handler(content_types=['document'])
def download_price(message):
    if not check_rights(message.from_user.id):
        replay_not_right_msg(message)
        return
    if command.current_state(message.from_user.id) != Commands.DOWNLOAD_PRICE:
        bot.reply_to(message, 'Если необходимо загрузить прайс, воспользуйтесь командой /download')
        return

    upload_file(message.document.file_id, 'price.xlsx')
    price_list.load_price('price.xlsx')
    command.clear(message.from_user.id)
    print(message)
    bot.reply_to(message, 'Прайс загружен!')


# Обработчик текстовых сообщений
@bot.message_handler(func=lambda message: True)
def echo_message(message):
    if not check_rights(message.from_user.id):
        replay_not_right_msg(message)
        return
    # Load your API key from an environment variable or secret management service
    print(f'request: {message.text}')
    result = price_list.search(message.text)
    response_txt = prepare_search_result(result)
    print(f"response: {response_txt}")
    bot.send_message(message.chat.id, f"Найдено {len(list(result.keys()))} позиций")
    bot.send_message(message.chat.id, response_txt, parse_mode='Markdown')


def prepare_search_result(result):
    print(result)
    table = pt.PrettyTable(['название', 'цена', 'остаток'])
    table.align['название'] = 'l'
    table.align['цена'] = 'r'
    table.align['остаток'] = 'r'
    for title, prices in result.items():
        table.add_row((title, prices[0].price, prices[0].rest))
    # response_txt = ''  # find_nomenclature(message.text)
    response_txt = f'```{table}```'
    print(f"response: {response_txt}")
    return response_txt


@bot.message_handler(content_types=['voice'])
def handle_docs_audio(message):
    if not check_rights(message.from_user.id):
        replay_not_right_msg(message)
        return
    # print(message)
    try:
        bot.send_message(message.chat.id, "Разбор голосового сообщения..")
        mp3_file = upload_voice(message.from_user.id, message.voice.file_id)

        with open(mp3_file, "rb") as audio_file:
            transcript = openai.Audio.transcribe(
                file=audio_file,
                model=Config.MODEL_VOICE_TRANSLIT,
                response_format="text",
                language="ru"
            )
        bot.send_message(message.chat.id, f"Ваше голосовое сообщение: {transcript}")
        clear_uploads(message.from_user.id)
        result = price_list.search(transcript.strip())
        response_txt = prepare_search_result(result)
        bot.send_message(message.chat.id, f"Найдено {len(list(result.keys()))} позиций")
        bot.send_message(message.chat.id, response_txt, parse_mode='Markdown')

    except Exception as exc:
        print(exc)
        bot.send_message(message.chat.id, '[!] error - {ошибка при загрузке голосового сообщения}')
        raise


def upload_voice(user_id, file_id):
    file_name = str(user_id) + '.oga'
    _file = upload_file(file_id, file_name)
    return save_oga_to_mp3(_file)


def upload_file(file_id, file_name):
    save_dir = os.getcwd()
    file_info = bot.get_file(file_id)
    downloaded_file = bot.download_file(file_info.file_path)

    _file = save_dir + "/upload/" + file_name
    print(_file)
    with open(_file, 'wb') as new_file:
        new_file.write(downloaded_file)
    return _file


def clear_uploads(user_id):
    save_dir = os.getcwd() + "/upload/"
    for file in glob.glob(save_dir + f"*{user_id}*.*"):
        os.unlink(file)


def check_rights(user_id):
    allow_users = json.loads(Config.ALLOW_USER_LIST)
    if int(user_id) not in allow_users:
        print(f'User blocked: {user_id}')
        return False
    return True


def replay_not_right_msg(message):
    bot.reply_to(
        message,
        f'У вас нет доступа, обратитесь к системному администратору! Ваш ID - {message.from_user.id}'
    )


# Запускаем бота
bot.polling()
# price_list.load_price('price.xlsx')
# print(price_list.search('all'))
